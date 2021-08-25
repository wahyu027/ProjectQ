from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep, strftime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os
import os.path
import sys
import re
import inc.kelas
import datetime


options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
driver = webdriver.Chrome(
    executable_path='chrome/chromedriver.exe', options=options)

driver.get("https://shopee.co.id")

try:
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable(
            (By.CLASS_NAME, 'shopee-popup__close-btn'))
    )

    popup = driver.find_element_by_class_name('shopee-popup__close-btn')
    popup.click()

    with open('tmp/daftar.html', 'w', encoding='utf-8') as file:
        file.write(driver.page_source)
        file.close()

except TimeoutException:
    print('Loading terlalu lama...')

finally:
    driver.quit()

if os.path.exists('tmp/daftar.html'):
    with open('tmp/daftar.html', 'r') as file:
        data = file.read()

    links = re.findall(
        'href="/top_products\?catId=(.*?)">.*?<div.*?">.*?<div.*?"></div>.*?<div.*?">.*?<img.*?">.*?</div>.*?<div.*?">.*?</div>.*?</div>.*?<div.*?">(.*?)</div>.*?</a>', data)

    wb = load_workbook(filename='template/daftar_produk.xlsx')
    ws = wb['daftar_produk']

    app = inc.kelas.Shopee()

    clean_items = []

    for index in range(len(links)):
        cat_id = links[index][0]

        print('  -> Mengambil data kategori ke ' +
              str(index + 1) + ' dari ' + str(len(links)) + ' : ' + links[index][1])

        top_product = app.ambil_terlaris(cat_id)

        top_item = top_product['data']['sections'][0]['data']['top_product'][0]['list']['data']['item']

        max_row = ws.max_row

        b = max_row + 1

        for one_item in top_item:
            one_cat_name = links[index][1]
            itemid = one_item['itemid']
            shopid = one_item['shopid']
            name = one_item['name']
            shop_name = one_item['shop_name']
            price = one_item['price_min']
            sold = one_item['sold']
            historical_sold = one_item['historical_sold']
            stock = one_item['stock']

            if itemid in clean_items:
                pass
            else:
                clean_items.append(itemid)

                ws['A' + str(b)] = str(b - 1)
                ws['A' + str(b)].alignment = Alignment(horizontal='center')
                ws['B' + str(b)] = one_cat_name
                ws['C' + str(b)] = str(itemid)
                ws['D' + str(b)] = str(shopid)
                ws['E' + str(b)] = name
                ws['F' + str(b)] = shop_name
                ws['G' + str(b)] = int(price) / 100000
                ws['G' + str(b)].number_format = '#,##0'
                ws['H' + str(b)] = int(sold)
                ws['H' + str(b)].number_format = '#,##0'
                ws['I' + str(b)] = int(historical_sold)
                ws['I' + str(b)].number_format = '#,##0'
                ws['J' + str(b)] = int(stock)
                ws['J' + str(b)].number_format = '#,##0'

                b = b + 1

        sleep(2)

    waktu = datetime.datetime.now()
    wb.save(filename=strftime('xlsx/produk_terlaris_' + '%d%m%Y_%H%M%S') + '.xlsx')

    os.remove('tmp/daftar.html')

    print('')
    print('  Proses selesai...')
    print('  File bisa dibuka di folder xlsx...')
    print('  Terima kasih...')
    print('')


else:
    print('File daftar.html tidak ditemukan...')
    sys.exit('Aplikasi berhenti...')
